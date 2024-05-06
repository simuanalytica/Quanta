import pandas as pd
import json
from scipy.stats import lognorm, norm, expon
from openpyxl import load_workbook
import os
import tkinter as tk
def parse_log_file(log_file_path):
    with open(log_file_path, 'r') as file:
        log_data = json.load(file)

    distributions = []

    for key, value in log_data.items():
        if 'distribution' in value:
            distribution_data = value['distribution']

            if 'lognorm' in distribution_data:
                distribution_type = 'lognorm'
                params = distribution_data['lognorm']
                distributions.append((key, distribution_type, params))

            elif 'norm' in distribution_data:
                distribution_type = 'norm'
                params = distribution_data['norm']
                distributions.append((key, distribution_type, params))

            elif 'expon' in distribution_data:
                distribution_type = 'expon'
                params = distribution_data['expon']
                distributions.append((key, distribution_type, params))

    return distributions

def generate_samples(distribution_type, params, num_samples):
    # print(num_samples)
    # # Convert parameters to numeric format
    s = float(params.get('s', 1.0))
    # print(s)
    loc = float(params['loc'])
    # print(loc)
    scale = float(params['scale'])
    # print(scale)
    # print(type(num_samples))
    # # Convert num_samples to an integer
    num_samples = int(num_samples)
    if distribution_type == 'lognorm':
        samples = lognorm.rvs(s=s, loc=loc, scale=scale, size=num_samples)
    elif distribution_type == 'norm':
        samples = norm.rvs(loc=loc, scale=scale, size=num_samples)
    elif distribution_type == 'expon':
        samples = expon.rvs(loc=loc, scale=scale, size=num_samples)
    else:
        raise ValueError(f"Unsupported distribution type: {distribution_type}")

    return samples

def generate_model(self):
    log_file_path = self.entry_dist_file.get()
    # Read distribution information from the JSON file
    with open(log_file_path, 'r') as json_file:
        json_data = json.load(json_file)

    num_samples = int(self.entry_iteration.get())  # Convert num_samples to an integer
    model_selected = self.combo_box.get()

    if model_selected.lower() == 'agile':
        model_file = 'Agile.xlsm'
    else:
        model_file = 'Default.xlsm'

    # Extract the directory path from the log file path
    log_directory = os.path.dirname(log_file_path)

    # Construct the path for the model file in the same directory as the log file
    model_file_path = os.path.join(log_directory, model_file)

    distributions = parse_log_file(log_file_path)
    data = {}

    for key, distribution_type, params in distributions:
        samples = generate_samples(distribution_type, params, num_samples)
        data[key] = samples

    # Create a DataFrame from the generated samples
    df = pd.DataFrame(data)

    # Load existing workbook based on the selected model
    workbook = load_workbook(model_file, keep_vba=True)

    # Get the Samples sheet
    samples_sheet = workbook['Samples']
    # Get the Input sheet
    input_sheet = workbook['Input']

    # Write DataFrame headers to the Samples sheet starting from cell A1
    for c_idx, header in enumerate(df.columns, start=1):
        samples_sheet.cell(row=1, column=c_idx, value=header)

    # Write DataFrame values to the Samples sheet starting from cell A2
    for r_idx, row in enumerate(df.iterrows(), start=2):
        for c_idx, value in enumerate(row[1], start=1):
            samples_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Define the starting row and column indices
    start_row = 5
    start_column = 4  # Column D

    # Updating the Input Sheet
    for row, (param_name, param_data) in enumerate(json_data.items(), start=start_row):
        input_sheet.cell(row=row, column=start_column).value = param_name

    # Save the workbook as .xlsm
    workbook.save(model_file_path.replace('.xlsx', '.xlsm'))
    output_text = f"Model Generation is completed.\n"
    self.output_text.insert(tk.END, output_text)

